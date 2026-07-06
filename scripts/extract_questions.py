#!/usr/bin/env python3
"""Extract exam question banks into a browser-friendly JS data file."""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE_DIR = Path(
    "/Users/higher/Library/CloudStorage/OneDrive-个人/02个人文档资料/01学习资料/"
    "交易员考试/交易员考试"
)
SOURCES = {
    "高级工": SOURCE_DIR / "电力交易员（高级工）题库.xlsx",
    "技师": SOURCE_DIR / "电力交易员（技师）题库.xlsx",
}
OUTPUT = Path(__file__).resolve().parents[1] / "data" / "questions.js"
LETTERS = "ABCDEFGHIJ"


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact(value: object) -> str:
    return re.sub(r"\s+", "", clean(value))


def normalize_answer(value: object, qtype: str) -> list[str]:
    text = compact(value)
    if qtype == "判断":
        return [text.replace("正确", "对").replace("错误", "错")]
    return [letter for letter in LETTERS if letter in text.upper()]


def question_scope(levels: list[str]) -> str:
    if levels == ["高级工"]:
        return "高级工独有"
    if levels == ["技师"]:
        return "技师新增"
    return "高级工+技师"


def extract_bank(level: str, source: Path) -> list[dict[str, object]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    questions = []

    for sheet in workbook.worksheets:
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            stem = clean(row[0] if len(row) > 0 else "")
            qtype = clean(row[1] if len(row) > 1 else "")
            answer = clean(row[2] if len(row) > 2 else "")
            if not stem or not qtype or not answer:
                continue

            options = []
            if qtype in {"单选", "多选"}:
                for idx, value in enumerate(row[3:13]):
                    label = LETTERS[idx]
                    text = clean(value)
                    if text:
                        options.append({"label": label, "text": text})
            elif qtype == "判断":
                options = [{"label": "对", "text": "对"}, {"label": "错", "text": "错"}]

            questions.append(
                {
                    "level": level,
                    "id": f"{level}-{sheet.title}-{row_number}",
                    "type": qtype,
                    "stem": stem,
                    "options": options,
                    "answer": normalize_answer(answer, qtype),
                }
            )

    return questions


def exact_key(question: dict[str, object]) -> tuple[object, ...]:
    options = tuple((option["label"], compact(option["text"])) for option in question["options"])  # type: ignore[index]
    return (
        compact(question["stem"]),
        question["type"],
        tuple(question["answer"]),  # type: ignore[arg-type]
        options,
    )


def extract() -> dict[str, object]:
    raw_questions = []
    source_counts: dict[str, int] = {}
    for level, source in SOURCES.items():
        bank_questions = extract_bank(level, source)
        raw_questions.extend(bank_questions)
        source_counts[level] = len(bank_questions)

    merged: dict[tuple[object, ...], dict[str, object]] = {}
    for question in raw_questions:
        key = exact_key(question)
        if key not in merged:
            merged[key] = {
                "stem": question["stem"],
                "type": question["type"],
                "options": question["options"],
                "answer": question["answer"],
                "levels": [],
            }

        item = merged[key]
        level = str(question["level"])
        if level not in item["levels"]:  # type: ignore[operator]
            item["levels"].append(level)  # type: ignore[index,union-attr]

    level_order = {"高级工": 0, "技师": 1}
    questions: list[dict[str, object]] = []
    for index, item in enumerate(merged.values(), start=1):
        levels = sorted(item["levels"], key=lambda value: level_order[str(value)])  # type: ignore[arg-type]
        item["levels"] = levels
        item["scope"] = question_scope([str(level) for level in levels])
        item["id"] = f"Q{index:04d}"
        questions.append(item)

    by_type: dict[str, int] = {}
    by_scope: dict[str, int] = {}
    for question in questions:
        qtype = str(question["type"])
        by_type[qtype] = by_type.get(qtype, 0) + 1
        scope = str(question["scope"])
        by_scope[scope] = by_scope.get(scope, 0) + 1

    return {
        "meta": {
            "title": "电力交易员高级工+技师题库",
            "sourceFiles": [source.name for source in SOURCES.values()],
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "rawTotal": len(raw_questions),
            "total": len(questions),
            "deduped": len(raw_questions) - len(questions),
            "sourceCounts": source_counts,
            "byType": by_type,
            "byScope": by_scope,
        },
        "questions": questions,
    }


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    data = extract()
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    OUTPUT.write_text(f"window.QUESTION_BANK = {payload};\n", encoding="utf-8")
    meta = data["meta"]
    print(f"Wrote {OUTPUT}")
    print(f"Total: {meta['total']} | By type: {meta['byType']}")


if __name__ == "__main__":
    main()
